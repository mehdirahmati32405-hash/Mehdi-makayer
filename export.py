"""ساخت فایل اکسل خروجی از گزارش‌ها — جدا از handlers.py تا بدون نیاز به
aiogram هم قابل تست باشد."""

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

FONT_NAME = "DejaVu Sans"
NAVY = "1F4E5F"
WHITE = "FFFFFF"
LIGHT = "F2F6F7"

HEADERS = [
    "ردیف", "کد پیگیری", "تاریخ ثبت", "دسته‌بندی", "توضیحات",
    "وضعیت", "بازخورد شهروند", "ناشناس؟", "از طرف", "موقعیت (lat,lon)",
]


def _report_code(report_id: int) -> str:
    return f"ML-{report_id:04d}"


def build_report_workbook(reports) -> bytes:
    """reports: لیستی از sqlite3.Row یا dict با همان کلیدهای جدول reports"""
    wb = Workbook()
    ws = wb.active
    ws.title = "گزارش‌ها"
    ws.sheet_view.rightToLeft = True

    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(name=FONT_NAME, bold=True, color=WHITE, size=11)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True, readingOrder=2)
    ws.row_dimensions[1].height = 26

    for i, r in enumerate(reports, start=1):
        row_idx = i + 1
        is_anon = bool(r["is_anonymous"]) if r["is_anonymous"] is not None else False
        values = [
            i,
            _report_code(r["id"]),
            (r["created_at"] or "")[:19].replace("T", " "),
            r["category"] or "",
            r["description"] or "",
            r["status"] or "",
            r["feedback"] or "",
            "بله" if is_anon else "خیر",
            "ناشناس" if is_anon else (r["user_name"] or ""),
            f"{r['latitude']}, {r['longitude']}" if r["latitude"] is not None else "",
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(name=FONT_NAME, size=10)
            cell.alignment = Alignment(
                horizontal="right" if col_idx in (4, 5) else "center",
                vertical="center", wrap_text=col_idx in (4, 5), readingOrder=2,
            )
            if i % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=LIGHT)

    widths = [6, 12, 20, 16, 34, 14, 20, 8, 16, 20]
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
